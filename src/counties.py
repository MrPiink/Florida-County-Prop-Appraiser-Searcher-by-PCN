import time

# pip install selenium
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException

from openpyxl.styles import PatternFill  # pip install openpyxl
from constants import DRIVER_PATH, EXCEL_FILE_LOCATION, WB, WS


class Counties:
    def __init__(
        self,
        name,
        appraiser_url,
        row_start,
        row_end,
        log_file_path,
        success_log_file_path,
        search_css,
        choose_pcn_css=None,
        is_choose_pcn_search=False,
        is_select_property=False,
        property_result_css=None,
        is_agree_button=False,
        agree_button_css=None,
        is_brevard=False,
        is_broward=False,
        is_duval=False,
        is_hillsborough=False,
        is_lee=False,
        is_marion=False,
        is_miami_dade=False,
        is_orange=False,
        is_osceola=False,
        is_palmbeach=False,
        is_pasco=False,
        is_polk=False,
        is_saint_lucie=False,
        is_sarasota=False,
        is_volusia=False,
        is_by_class=False
    ):

        self.name = name
        self.appraiser_url = appraiser_url
        self.row_start = row_start
        self.row_end = row_end + 1
        self.log_file_path = log_file_path
        self.success_log_file_path = success_log_file_path
        self.search_css = search_css
        self.choose_pcn_css = choose_pcn_css
        self.is_choose_pcn_search = is_choose_pcn_search
        self.is_select_property = is_select_property
        self.property_result_css = property_result_css
        self.is_agree_button = is_agree_button
        self.agree_button_css = agree_button_css
        self.is_brevard = is_brevard
        self.is_broward = is_broward
        self.is_duval = is_duval
        self.is_hillsborough = is_hillsborough
        self.is_lee = is_lee
        self.is_marion = is_marion
        self.is_miami_dade = is_miami_dade
        self.is_orange = is_orange
        self.is_osceola = is_osceola
        self.is_palmbeach = is_palmbeach
        self.is_pasco = is_pasco
        self.is_polk = is_polk
        self.is_saint_lucie = is_saint_lucie
        self.is_sarasota = is_sarasota
        self.is_volusia = is_volusia
        self.is_by_class = is_by_class

    def get_pcn_num(self):
        """Goes row by row pulling each pcn and appending it to the dictionary with the row as its key."""
        pcn_nums_dict = {}
        for row in range(self.row_start, self.row_end):
            pcn_num = WS.cell(row, 16).value
            pcn_nums_dict[row] = pcn_num
        return pcn_nums_dict

    def choose_pcn_search(self, driver, owner_names_dict, row):
        """Uses selenium to detect the pcn search button on the website being searched. Uses css to locate the button and then click it."""
        try:
            pcn_button = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, self.choose_pcn_css))
            )
            pcn_button.click()
            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "BUTTON ERROR"
            return 0

    def search_pcn_num(self, driver, pcn_num, owner_names_dict, row):
        """Searches the pcn number for the given row in the search bar on the website being searched."""
        try:
            if self.is_polk and len(str(pcn_num)) > 18:
                driver.quit()
                owner_names_dict[row] = "SEARCH ERROR"
                return 0

            # If the counties corresponding property appraiser site's css does not work for searching specify is_by_class in the objects creation.
            if self.is_by_class:
                search = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located(
                        (By.CLASS_NAME, self.search_css))  # The self.seach_css varible here contains a class name instead.
                )

            else:
                search = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located(
                        (By.CSS_SELECTOR, self.search_css))
                )
            search.send_keys(pcn_num)
            search.send_keys(Keys.RETURN)
            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "SEARCH ERROR"
            return 0

    def search_pcn_num_duval(self, driver, pcn_num, owner_names_dict, row):
        """Duval county's website requires a different way of searching. This function is used instead of the previous search function."""
        pcn_num_list = pcn_num.split("-")
        try:
            search = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "#ctl00_cphBody_tbRE6"))
            )
            search.send_keys(pcn_num_list[0])

            search = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "#ctl00_cphBody_tbRE4"))
            )
            search.send_keys(pcn_num_list[1])
            search.send_keys(Keys.RETURN)
            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "SEARCH ERROR"
            return 0

    def search_pcn_num_pasco(self, driver, pcn_num, owner_names_dict, row):
        """Pasco county's website requires a different way of searching. This function is used instead of search_pcn_num"""

        pcn_num_list = pcn_num.split("-")
        if len("".join(pcn_num_list)) > 19:
            driver.quit()
            owner_names_dict[row] = "SEARCH ERROR"
            return 0

        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "#Form1 > table > tbody > tr.sentry > td:nth-child(2) > input[type=text]"))
            )
            for n, num in enumerate(pcn_num_list, start=2):
                search = driver.find_element(
                    By.CSS_SELECTOR, f"#Form1 > table > tbody > tr.sentry > td:nth-child({n}) > input[type=text]")
                search.send_keys(num)

            search.send_keys(Keys.RETURN)
            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "SEARCH ERROR"
            return 0

    def select_property(self, driver, owner_names_dict, row):
        """Some appraiser sites do not immediately go to the properties page and instead prompt you to select one even if there is only one. This clicks the first result."""
        try:
            property_result = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, self.property_result_css))
            )
            property_result.click()
            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "SELECTION ERROR"
            return 0

    def select_agree(self, driver):
        """In the event a website contains an agree button this clicks it."""
        try:
            agree_button = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, self.agree_button_css))
            )
            agree_button.click()

        except TimeoutException:
            pass

    # All get_owners_[county name] are the same but constructed to pull the owner names based off of the layout of the website.
    # driver.quit must come after owner_names.text.
    def get_owners_brevard(self, driver, owner_names_dict, row):
        try:
            owner_names = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "#cssDetails_Top_Outer > div.cssDetails_TopContainer.cssTableContainer.cssOverFlow_x.cssCanReceiveFocus > div > div:nth-child(1) > div.cssDetails_Top_Cell_Data > div:nth-child(1)")
                )
            )

            owner_names_dict[row] = owner_names.text
            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "NAME ERROR"
            return 0

    def get_owners_broward(self, driver, owner_names_dict, row):
        try:
            owner_names1 = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "#ownerNameId"))
            )
            owner_names2 = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "#ownerName2Id"))
            )

            if owner_names2.text != " ":
                owner_names_dict[row] = owner_names1.text + \
                    " " + owner_names2.text

            else:
                owner_names_dict[row] = owner_names1.text

            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "NAME ERROR"
            return 0

    def get_owners_duval(self, driver, owner_names_dict, row):
        owner_names_list = []
        n = 0

        try:
            owner_names = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR,
                     f"#ctl00_cphBody_repeaterOwnerInformation_ctl0{n}_lblOwnerName")
                )
            )
            owner_names_list.append(owner_names.text)

            while True:
                try:
                    n += 1
                    owner_names = driver.find_element(
                        By.CSS_SELECTOR, f"#ctl00_cphBody_repeaterOwnerInformation_ctl0{n}_lblOwnerName")
                    owner_names_list.append(owner_names.text)

                except NoSuchElementException:
                    break

            driver.quit()
            owner_names_dict[row] = " ".join(owner_names_list)
            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "NAME ERROR"
            return 0

    def get_owners_hillsborough(self, driver, owner_names_dict, row):
        try:
            owner_names = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "#details > div.property-data > div.propcard-section > h4"))
            )
            owner_names_dict[row] = owner_names.text.replace("\n", " ")
            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "NAME ERROR"
            return 0

    def get_owners_lee(self, driver, owner_names_dict, row):
        try:
            owner_names = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "#divDisplayParcelOwner > div.column.columnLeft > div > div.textPanel > div"))
            )

            owner_names_dict[row] = owner_names.text
            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "NAME ERROR"
            return 0

    def get_owners_marion(self, driver, owner_names_dict, row):
        try:
            owner_names = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "#prc > table > tbody > tr > td > table:nth-child(10) > tbody > tr > td:nth-child(1)"))
            )
            owner_names_dict[row] = owner_names.text
            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "NAME ERROR"
            return 0

    def get_owners_miami_dade(self, driver, owner_names_dict, row):
        owner_names_list = []
        n = 1

        try:
            owner_names = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR,
                     f"#property_info > tbody > tr:nth-child(5) > td > div > div:nth-child({n}) > div")
                )
            )
            owner_names_list.append(owner_names.text)

            while True:
                try:
                    n += 1
                    owner_names = driver.find_element(
                        By.CSS_SELECTOR, f"#property_info > tbody > tr:nth-child(5) > td > div > div:nth-child({n}) > div")
                    owner_names_list.append(owner_names.text)

                except NoSuchElementException:
                    break

            driver.quit()
            owner_names_dict[row] = " ".join(owner_names_list)
            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "NAME ERROR"
            return 0

    def get_owners_orange(self, driver, owner_names_dict, row):
        try:
            owner_names = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "#ngb-nav-5-panel > parcel-card-component > div.jumbotron.m-0.border > div:nth-child(1) > div.col > div:nth-child(2) > div:nth-child(1) > div > span.multiLine"))
            )
            owner_names_dict[row] = owner_names.text.replace("\n", "")
            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "NAME ERROR"
            return 0

    def get_owners_osceola(self, driver, owner_names_dict, row):
        try:
            owner_names = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "#owner-information-table > tbody > tr:nth-child(1) > td"))
            )
            owner_names_dict[row] = owner_names.text
            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "NAME ERROR"
            return 0

    def get_owners_palmbeach(self, driver, owner_names_dict, row):
        owner_names_list = []
        n = 2

        try:
            owner_names = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR,
                     f"#ownerInformationDiv > fieldset > table > tbody > tr:nth-child(2) > td:nth-child(1) > table > tbody > tr:nth-child({n}) > td")
                )
            )
            owner_names_list.append(owner_names.text)

            while True:
                try:
                    n += 1
                    owner_names = driver.find_element(
                        By.CSS_SELECTOR, f"#ownerInformationDiv > fieldset > table > tbody > tr:nth-child(2) > td:nth-child(1) > table > tbody > tr:nth-child({n}) > td")
                    owner_names_list.append(owner_names.text)

                except NoSuchElementException:
                    break

            driver.quit()
            owner_names_dict[row] = " ".join(owner_names_list)
            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "NAME ERROR"
            return 0

    def get_owners_pasco(self, driver, owner_names_dict, row):
        try:
            owner_names = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "#lblMailingAddress"))
            )
            owner_names_dict[row] = owner_names.text
            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "NAME ERROR"
            return 0

    def get_owners_polk(self, driver, owner_names_dict, row):
        owner_names_list = []
        n = 1

        try:
            owner_names = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR,
                     f"#CamaDisplayArea > table:nth-child(3) > tbody > tr > td:nth-child(1) > table:nth-child(2) > tbody > tr:nth-child({n})")
                )
            )
            owner_names_list.append(owner_names.text)

            while True:
                try:
                    n += 1
                    owner_names = driver.find_element(
                        By.CSS_SELECTOR, f"#CamaDisplayArea > table:nth-child(3) > tbody > tr > td:nth-child(1) > table:nth-child(2) > tbody > tr:nth-child({n})")
                    owner_names_list.append(owner_names.text)

                except NoSuchElementException:
                    break

            driver.quit()
            owner_names_dict[row] = " ".join(owner_names_list)
            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "NAME ERROR"
            return 0

    def get_owners_saint_lucie(self, driver, owner_names_dict, row):
        owner_names_list = []
        n = 1

        try:
            owner_names = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR,
                     f"body > div > div > div.container.propCard > div.ng-isolate-scope > div > div.tab-pane.ng-scope.active > div > div > div:nth-child(2) > div.col-sm-6.padme-bottom-10 > div:nth-child({n})")
                )
            )
            owner_names_list.append(owner_names.text)

            while True:
                try:
                    n += 1
                    owner_names = driver.find_element(
                        By.CSS_SELECTOR, f"body > div > div > div.container.propCard > div.ng-isolate-scope > div > div.tab-pane.ng-scope.active > div > div > div:nth-child(2) > div.col-sm-6.padme-bottom-10 > div:nth-child({n})")
                    owner_names_list.append(owner_names.text)

                except NoSuchElementException:
                    break

            driver.quit()
            owner_names_list = owner_names_list[0:len(owner_names_list)-1]
            owner_names_dict[row] = " ".join(owner_names_list)
            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "NAME ERROR"
            return 0

    def get_owners_sarasota(self, driver, owner_names_dict, row):
        owner_names_list = []
        n = 2

        try:
            owner_names = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR,
                     f"#container > ul.resultl.spaced > li:nth-child({n})")
                )
            )
            owner_names_list.append(owner_names.text)

            while True:
                try:
                    n += 1
                    owner_names = driver.find_element(
                        By.CSS_SELECTOR, f"#container > ul.resultl.spaced > li:nth-child({n})")
                    owner_names_list.append(owner_names.text)

                except NoSuchElementException:
                    break

            driver.quit()
            owner_names_list = owner_names_list[0:len(owner_names_list)-4]
            owner_names_dict[row] = " ".join(owner_names_list)
            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "NAME ERROR"
            return 0

    def get_owners_volusia(self, driver, owner_names_dict, row):
        try:
            owner_names = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "body > main > div:nth-child(8) > div:nth-child(1) > div > div:nth-child(14)"))
            )
            owner_names_dict[row] = owner_names.text
            return 1

        except TimeoutException:
            driver.quit()
            owner_names_dict[row] = "NAME ERROR"
            return 0

    def get_owner_names(self):
        """This function triggers all of the website related functions to navigate to and pull the owner name for each pcn/row."""
        pcn_nums_dict = self.get_pcn_num()
        owner_names_dict = {}

        for row, pcn_num in pcn_nums_dict.items():
            options = webdriver.ChromeOptions()
            # Removes constant webdriver logging.
            options.add_argument("--log-level=3")
            if not self.is_orange:
                # Orange county cannot be run in headless mode.
                options.add_argument("headless")
            driver = webdriver.Chrome(DRIVER_PATH, options=options)
            driver.get(self.appraiser_url)
            # Lee and Saint Lucie county open new tabs during website navigation. This gets the current tab.
            if self.is_lee or self.is_saint_lucie:
                current_tab = driver.window_handles

            if self.is_agree_button:
                self.select_agree(driver)

            if self.is_choose_pcn_search:
                result = self.choose_pcn_search(driver, owner_names_dict, row)

            # Duval and Pasco use separate search functions.
            if not self.is_choose_pcn_search or result == 1:
                if self.is_duval:
                    result = self.search_pcn_num_duval(
                        driver, pcn_num, owner_names_dict, row
                    )

                elif self.is_pasco:
                    result = self.search_pcn_num_pasco(
                        driver, pcn_num, owner_names_dict, row)

                else:
                    result = self.search_pcn_num(
                        driver, pcn_num, owner_names_dict, row)

            if self.is_select_property and result == 1:
                if self.is_saint_lucie:
                    # Saint Lucie County always has results showing this makes sure the searched result loads before selecting the first one.
                    time.sleep(3)
                result = self.select_property(driver, owner_names_dict, row)

            # Wait for property page to fully load before extracting owner names.
            time.sleep(3)

            if self.is_brevard and result == 1:
                result = self.get_owners_brevard(driver, owner_names_dict, row)

            elif self.is_broward and result == 1:
                result = self.get_owners_broward(driver, owner_names_dict, row)

            elif self.is_duval and result == 1:
                result = self.get_owners_duval(driver, owner_names_dict, row)

            elif self.is_hillsborough and result == 1:
                result = self.get_owners_hillsborough(
                    driver, owner_names_dict, row)

            elif self.is_lee and result == 1:
                # This switches to the new tab.
                tab_list = driver.window_handles
                new_tab = list(set(tab_list) - set(current_tab))[0]
                driver.switch_to.window(new_tab)
                time.sleep(1)
                try:
                    # The agree button only comes up once every few hours. If the button is not there pass.
                    self.select_agree(driver)
                except Exception:
                    pass
                result = self.get_owners_lee(driver, owner_names_dict, row)

            elif self.is_marion and result == 1:
                result = self.get_owners_marion(driver, owner_names_dict, row)

            elif self.is_miami_dade and result == 1:
                result = self.get_owners_miami_dade(
                    driver, owner_names_dict, row)

            elif self.is_orange and result == 1:
                result = self.get_owners_orange(
                    driver, owner_names_dict, row)

            elif self.is_osceola and result ==1:
                result = self.get_owners_osceola(
                    driver, owner_names_dict, row)

            elif self.is_palmbeach and result == 1:
                result = self.get_owners_palmbeach(
                    driver, owner_names_dict, row
                )

            elif self.is_pasco and result == 1:
                result = self.get_owners_pasco(
                    driver, owner_names_dict, row
                )

            elif self.is_polk and result == 1:
                result = self.get_owners_polk(
                    driver, owner_names_dict, row
                )

            elif self.is_saint_lucie and result == 1:
                # This switches to the new tab.
                tab_list = driver.window_handles
                new_tab = list(set(tab_list) - set(current_tab))[0]
                driver.switch_to.window(new_tab)
                time.sleep(1)
                result = self.get_owners_saint_lucie(
                    driver, owner_names_dict, row)

            elif self.is_sarasota and result == 1:
                result = self.get_owners_sarasota(
                    driver, owner_names_dict, row
                )

            elif self.is_volusia and result == 1:
                result = self.get_owners_volusia(
                    driver, owner_names_dict, row
                )

            # Prints the status of the program after each row.
            print(f"{self.name}: {row}/{self.row_end - 1}")
        return owner_names_dict

    def format_owner_names(self):
        """Formats owner names pulled from website to be upper case and have no new line characters."""
        owner_names_dict = self.get_owner_names()
        f_owner_names_dict = {}
        for row, owner_names in owner_names_dict.items():
            if "\n" in owner_names:
                owner_names = owner_names.replace("\n", " ")
            owner_names = owner_names.upper()
            f_owner_names_dict[row] = owner_names
        return f_owner_names_dict

    @ staticmethod
    def get_debtor_fn(row):
        """Pulls the debtors first name from the spreadsheet for the corresponding row."""
        fn_cell = WS.cell(row, 4)
        debtor_first_name = fn_cell.value
        if debtor_first_name is None:
            return "N/A"
        debtor_first_name_separated = debtor_first_name.split(" ")
        debtor_first_name = debtor_first_name_separated[0]
        return debtor_first_name.upper()

    @ staticmethod
    def get_debtor_ln(row):
        """Pulls the debtors last name from the spreadsheet for the corresponding row."""
        ln_cell = WS.cell(row, 3)
        debtor_last_name = ln_cell.value
        if debtor_last_name is None:
            return "N/A"
        return debtor_last_name.upper()

    def logger(self, log_type, debtor_first_name, debtor_last_name, row, owner_names="none"):
        """Logs the outcome of the program to a logfile and logfile that just contains matches for each pcn searched."""
        if log_type == "SUCCESSFUL MATCH":
            success_log_file = open(self.success_log_file_path, "a")
            success_log_file.write(
                f"{log_type} | {owner_names} | {debtor_first_name} {debtor_last_name} | ROW: {row}\n")
            success_log_file.close()

        log_file = open(self.log_file_path, "a")
        log_file.write(
            f"{log_type} | {owner_names} | {debtor_first_name} {debtor_last_name} | ROW: {row}\n"
        )
        log_file.close()

    @ staticmethod
    def excel_cell_highlighter(color_hex, row):
        """Highlights the first cell of each row depending on the outcome of the program."""
        WS.cell(row, 1).fill = PatternFill(
            start_color=color_hex, fill_type="solid")
        WB.save(filename=EXCEL_FILE_LOCATION)

    def compare_owner_names(self):
        """The final function that compares the names pulled from the property appraiser sites to the debtor names on the spreadsheet. 
        For each row the debtor name is compared to the name pulled based on the pcn search. 
        The search is generous and will consider it a match even if only the first or last name match.
        This is required since the property appraiser website might not be displaying the owner name perfectly."""
        f_owner_names_dict = self.format_owner_names()
        for row, owner_names in f_owner_names_dict.items():
            debtor_first_name = self.get_debtor_fn(row)
            debtor_last_name = self.get_debtor_ln(row)

            if owner_names == "AGREE ERROR":
                self.logger("AGREE ERROR", debtor_first_name,
                            debtor_last_name, row)

            elif owner_names == "BUTTON ERROR":
                self.logger("BUTTON ERROR", debtor_first_name,
                            debtor_last_name, row)

            elif owner_names == "SEARCH ERROR":
                self.logger("SEARCH ERROR", debtor_first_name,
                            debtor_last_name, row)

            elif owner_names == "SELECTION ERROR":
                self.logger("SELECTION ERROR", debtor_first_name,
                            debtor_last_name, row)

            elif owner_names == "NAME ERROR":
                self.logger("NAME ERROR", debtor_first_name,
                            debtor_last_name, row)

            else:
                if debtor_first_name in owner_names or debtor_last_name in owner_names:
                    self.logger("SUCCESSFUL MATCH", debtor_first_name,
                                debtor_last_name, row, owner_names=owner_names)
                    self.excel_cell_highlighter(
                        "00FF00", row)  # Highlights row green

                else:
                    self.logger("FAILED TO MATCH", debtor_first_name,
                                debtor_last_name, row, owner_names=owner_names)
                    self.excel_cell_highlighter(
                        "FF0000", row)  # Highlights row red
        # Prints when the program has finished the county.
        print(f"{self.name}: Done!")

    def start(self):
        """Initiates the searching for the county object."""
        self.compare_owner_names()
